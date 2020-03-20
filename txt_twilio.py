from twilio.rest import Client
import os
import openpyxl

#read in spreadsheet with data on revenue by product, or any other data source
wb = openpyxl.load_workbook('Revenues.xlsx')

#indicate what sheet you are interested in 
sheet = wb.get_sheet_by_name('Sheet1')

#find the last column that contains revenues
lastCol = sheet.max_column

#create empty dictionary with information on items that raeched revenue goal

reachedGoals = {}

#loop through each row and add to the dictionary only items that reached the revenue goal

for r in range(2, sheet.max_row + 1):
  revenue = sheet.cell(row=r, column=lastCol).value
  if revenue > 80000:
    item = sheet.cell(row=r, column=1).value
    revenue = sheet.cell(row=r, column=2).value
    reachedGoals[item] = revenue

#Retrieve the TWILIO SID and TOKEN, the TWILIO NUMBER and the desired number you want 
#to send a text message to. You can get those by registering at www.twilio.com.
#Since those data are sensitive, I have assigned them to an environment variable.
#Below I retrieve each environment variable and assign it to an appropriate variable name.
#Learn more about enviroment variables here: https://able.bio/rhett/how-to-set-and-get-environment-variables-in-python--274rgt5 

accountSID = os.getenv('TWILIO_SID')
authToken  = os.getenv('TWILIO_TOKEN')
twilioCli = Client(accountSID, authToken)
myTwilioNumber = os.getenv('TWILIO_NUM')
myCellPhone = os.getenv('MY_NUM')

#I then check if there are at all any items in the data source that reached the goal
# If so, it will send a text message to my phone number with the list of items.
if len(reachedGoals)>0:
  message = twilioCli.messages.create(body=f'Hey Olga, \nhere are the items that reached revenue goal: \n{reachedGoals}', from_=myTwilioNumber, to=myCellPhone)
